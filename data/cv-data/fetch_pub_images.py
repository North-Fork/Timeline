"""Build publication cover image lookup from jasonlewis.org media library.

Strategy:
  1. Start with the manually curated MANUAL_MAP (website title → image URL).
  2. Fetch the full WP media library via REST API; for any website title not yet
     in the map, attempt keyword-based matching against new media items.
     This auto-picks up new cover images uploaded to jasonlewis.org.
  3. Read cv.xlsx; for Dissemination entries, fuzzy-match cv headlines to website
     titles; record the image URL.
  4. Write pub-images.js (window.__PUB_IMAGES__ = { cvHeadline: imageUrl, ... }).

Note: jasonlewis.org/category/publication/ is JavaScript-rendered and cannot be
scraped via a static HTTP fetch. New titles should be added to WEBSITE_TITLES
manually when they appear on the site; new cover images are auto-discovered via
the WP media REST API (step 2).

Usage:
    cd data/cv-data
    python3 fetch_pub_images.py

Output: pub-images.js
"""

import re
import json
import urllib.request
import urllib.error
from pathlib import Path
import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────

WP_MEDIA_API = 'https://jasonlewis.org/wp-json/wp/v2/media'
XLSX_PATH    = Path(__file__).parent / 'cv.xlsx'
OUT_PATH     = Path(__file__).parent / 'pub-images.js'

DISSEMINATION_GROUPS = {
    'Books/Chapters', 'Books', 'Book Chapters',
    'Journal Articles',
    'Invited Publications',
    'Op-Ed',
}

# Known website publication titles (from jasonlewis.org/category/publication/).
# These are the keys used for matching; update list when new titles are added.
WEBSITE_TITLES = [
    "Curating Superintelligences",
    "Before Intelligence",
    "The Indigenous Protocol and AI Workshops as Future Imaginary",
    "CyberPowWow and the First Wave of Indigenous Media Arts",
    "Reworlding AI Through Future Imaginaries",
    "The Myths of My Descendants",
    "Building Aboriginal Territories in Cyberspace",
    "Abundant Intelligences: Placing AI within Indigenous Knowledge Frameworks",
    "Good Technology is Messy",
    "Expansive & Exuberant: The Future Imaginaries of Solomon Enos",
    "Making Kin with the Machines v.4",
    "Imagining Indigenous AI",
    "Relation-Oriented AI: Why Indigenous Protocols Matter for the Digital Humanities",
    "The Future Imaginary",
    "Overclock the Imagination! Mapping the Indigenous Future Imaginary",
    "Making Kin with the Machines v.3",
    "From Impoverished Intelligence to Abundant Intelligences",
    "Making Kin with the Machines v.2",
    "22nd-Century Proto:typing",
    "Indigenous Protocol and Artificial Intelligence Position Paper",
    "An Orderly Assemblage of Biases: Troubling the Monocultural Stack",
    "Future Imaginary Dialogue with Dr. Kim TallBear",
    "Making Kin with the Machines",
    "A Brief (Media) History of the Indigenous Future",
    "Preparations for a Haunting: Note Towards an Indigenous Future Imaginary",
    "A Better Dance and Better Prayers: Systems, Structures, and the Future Imaginary in Aboriginal New Media",
]

# Manually curated override map: website_title → source_url.
# Takes priority over any automated matching result.
# Built from WP media API (wp-json/wp/v2/media) — update as new covers are uploaded.
MANUAL_MAP = {
    "Curating Superintelligences":
        "https://jasonlewis.org/wp-content/uploads/2026/03/Curating-Superintelligence_frontcover.jpeg",
    "The Myths of My Descendants":
        "https://jasonlewis.org/wp-content/uploads/2025/04/The-Myths-of-My-Descendants.jpg",
    "Building Aboriginal Territories in Cyberspace":
        "https://jasonlewis.org/wp-content/uploads/2025/04/Building-Aboriginal-Territories-in-Cyberspace.jpg",
    "Abundant Intelligences: Placing AI within Indigenous Knowledge Frameworks":
        "https://jasonlewis.org/wp-content/uploads/2025/04/abundant-intelligences-thumbnail.jpg",
    "Good Technology is Messy":
        "https://jasonlewis.org/wp-content/uploads/2024/04/JEL-site-pubs_good-tech-is-messy.jpg",
    "Making Kin with the Machines v.4":
        "https://jasonlewis.org/wp-content/uploads/2024/04/JEL-site-pubs_making-kin-v4.jpg",
    "The Future Imaginary":
        "https://jasonlewis.org/wp-content/uploads/2024/04/JEL-Site-Pubs-RoutledgeHandbook-ofCoFuturisms-e1713211924499.png",
    "Imagining Indigenous AI":
        "https://jasonlewis.org/wp-content/uploads/2023/05/JEL-Site-Pubs-Imagining-AI-e1685463716472.jpeg",
    "Relation-Oriented AI: Why Indigenous Protocols Matter for the Digital Humanities":
        "https://jasonlewis.org/wp-content/uploads/2023/05/JEL-Site-Pubs-relation-oriented-AI-1-e1685471818651.jpeg",
    "Overclock the Imagination! Mapping the Indigenous Future Imaginary":
        "https://jasonlewis.org/wp-content/uploads/2023/05/JEL-Site-Pubs-Overclock-Imagination-e1685463996861.png",
    "Making Kin with the Machines v.2":
        "https://jasonlewis.org/wp-content/uploads/2021/11/JL-Making-Kin-with-The-Machines-v2.jpg",
    "Making Kin with the Machines":
        "https://jasonlewis.org/wp-content/uploads/2021/11/JL-Making-Kin-with-The-Machines.jpg",
    "22nd-Century Proto:typing":
        "https://jasonlewis.org/wp-content/uploads/2021/11/JL-22nd-Century-Proto-typing.jpg",
    "Indigenous Protocol and Artificial Intelligence Position Paper":
        "https://jasonlewis.org/wp-content/uploads/2024/10/Indigenous-Protocol-and-AI.-2020.-PUBLISHED_-1.png",
    "An Orderly Assemblage of Biases: Troubling the Monocultural Stack":
        "https://jasonlewis.org/wp-content/uploads/2019/04/orderly.jpg",
    "A Brief (Media) History of the Indigenous Future":
        "https://jasonlewis.org/wp-content/uploads/2019/04/abreifmedia.jpg",
    "Preparations for a Haunting: Note Towards an Indigenous Future Imaginary":
        "https://jasonlewis.org/wp-content/uploads/2019/04/prepforhaunt.jpg",
    # Cv headline variants that don't auto-match cleanly
    "Overclock the Imagination! Mapping the Indigenous Future Imaginary":
        "https://jasonlewis.org/wp-content/uploads/2023/05/JEL-Site-Pubs-Overclock-Imagination-e1685463996861.png",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

STOPWORDS = {'the', 'a', 'an', 'and', 'or', 'of', 'in', 'to', 'for', 'with',
             'by', 'from', 'as', 'at', 'is', 'why', 'note', 'notes', 'dr', 'ii',
             'towards', 'toward', 'its', 'our', 'their', 'this', 'that'}

def tokenize(text: str) -> set:
    """Lowercase, strip punctuation, split, remove stopwords."""
    words = re.sub(r"[^\w\s]", " ", text.lower()).split()
    return {w for w in words if w not in STOPWORDS and len(w) > 2}


def extract_quoted_title(headline: str) -> str | None:
    """Extract title from quoted citations — handles ASCII and Unicode curly quotes."""
    # Unicode curly quotes: \u201c (") and \u201d (")
    for pat in [r'\u201c([^\u201d]+)\u201d', r'"([^"]+)"']:
        m = re.search(pat, headline)
        if m:
            return m.group(1).rstrip('.,;').strip()
    return None


def best_match(cv_headline: str, website_image_map: dict) -> tuple[str | None, float]:
    """Return (website_title, score) for the best fuzzy match, or (None, 0)."""
    # Try extracted quoted title first, then full headline
    candidates = []
    quoted = extract_quoted_title(cv_headline)
    if quoted:
        candidates.append(quoted)
    candidates.append(cv_headline)

    best_title, best_score = None, 0.0
    # Sort site titles by token count descending: prefer longer, more specific matches
    sorted_titles = sorted(website_image_map.keys(),
                           key=lambda t: len(tokenize(t)), reverse=True)
    for i, candidate in enumerate(candidates):
        is_quoted = (i == 0 and quoted is not None)  # first candidate is extracted title
        cv_tokens = tokenize(candidate)
        if not cv_tokens:
            continue
        for site_title in sorted_titles:
            site_tokens = tokenize(site_title)
            if not site_tokens:
                continue
            overlap = len(cv_tokens & site_tokens)
            precision = overlap / len(cv_tokens)
            recall    = overlap / len(site_tokens)
            if precision + recall == 0:
                continue
            f1 = 2 * precision * recall / (precision + recall)
            # When using the full citation (not extracted title), skip cases where
            # the site title is a very short phrase fully contained in a much longer cv title
            # (prevents "The Future Imaginary" matching citations that just mention it)
            if not is_quoted and recall == 1.0 and precision < 0.3:
                continue
            if f1 > best_score:
                best_score, best_title = f1, site_title

    return (best_title, best_score) if best_score >= 0.50 else (None, 0.0)


def fetch_all_media() -> list[dict]:
    """Fetch all images from the WP media API."""
    items = []
    page = 1
    while True:
        url = f"{WP_MEDIA_API}?per_page=100&page={page}&media_type=image&_fields=id,slug,source_url,title"
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        try:
            with urllib.request.urlopen(req, timeout=10) as resp:
                batch = json.loads(resp.read().decode('utf-8'))
                if not batch:
                    break
                items.extend(batch)
                if len(batch) < 100:
                    break
                page += 1
        except Exception as e:
            print(f"  Warning: media API fetch failed (page {page}): {e}")
            break
    return items


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    # Step 1: Start with manual curated map
    print("=== Step 1: Using curated manual map ===")
    website_image_map = dict(MANUAL_MAP)
    print(f"  {len(website_image_map)} manually curated entries.")

    # Step 2: Try to fill gaps by fetching WP media API
    # New cover images uploaded to jasonlewis.org are auto-discovered here.
    # To add a new publication title, append it to WEBSITE_TITLES above.
    missing = [t for t in WEBSITE_TITLES if t not in website_image_map]
    if missing:
        print(f"\n=== Step 2: Fetching WP media API for {len(missing)} unmatched titles ===")
        try:
            all_media = fetch_all_media()
            print(f"  Fetched {len(all_media)} media items from WP.")

            for title in missing:
                title_tokens = tokenize(title)
                best_item, best_score = None, 0.0
                for item in all_media:
                    slug  = item.get('slug', '')
                    label = item.get('title', {})
                    label = label.get('rendered', slug) if isinstance(label, dict) else str(label)
                    media_tokens = tokenize(slug + ' ' + label)
                    overlap  = len(title_tokens & media_tokens)
                    coverage = overlap / len(title_tokens) if title_tokens else 0
                    if coverage > best_score:
                        best_score, best_item = coverage, item
                if best_item and best_score >= 0.4:
                    url = best_item.get('source_url', '')
                    website_image_map[title] = url
                    print(f"  ✓  {title[:50]:50s} (score={best_score:.2f}) → {url[:60]}")
                else:
                    print(f"  ✗  {title[:50]:50s} — no confident match")
        except Exception as e:
            print(f"  Warning: could not fetch WP media API: {e}")

    # Step 3: Read cv.xlsx and match Dissemination headlines
    print(f"\n=== Step 3: Matching cv.xlsx Dissemination entries ===")
    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found. Run generate_cv_xlsx.py first.")
        return

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active
    headers = [
        str(c.value).strip().lower().replace(' ', '') if c.value else ''
        for c in ws[1]
    ]
    try:
        h_col = headers.index('headline')
        g_col = headers.index('group')
    except ValueError as e:
        print(f"ERROR: column not found: {e}")
        return

    pub_images = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= max(h_col, g_col):
            continue
        raw_headline = str(row[h_col]).strip() if row[h_col] else ''
        group        = str(row[g_col]).strip()  if row[g_col]  else ''
        if not raw_headline or group not in DISSEMINATION_GROUPS:
            continue

        site_title, score = best_match(raw_headline, website_image_map)
        if site_title:
            pub_images[raw_headline] = website_image_map[site_title]
            print(f"  ✓  {raw_headline[:48]:48s} → {site_title[:35]} ({score:.2f})")
        else:
            print(f"  ✗  {raw_headline[:48]}")

    # Step 4: Also add website titles as direct keys (for GDoc-loaded data
    #          where ev.headline matches the website title exactly)
    for title, img_url in website_image_map.items():
        if title not in pub_images:
            pub_images[title] = img_url

    # Step 5: Write pub-images.js
    js_data = json.dumps(pub_images, indent=2, ensure_ascii=False)
    cv_matched = sum(1 for k in pub_images if k not in website_image_map)
    OUT_PATH.write_text(
        f"// Auto-generated by fetch_pub_images.py — do not edit manually.\n"
        f"// Re-run after updating jasonlewis.org or cv.xlsx.\n"
        f"// {len(website_image_map)} website titles with images; "
        f"{cv_matched} additional cv.xlsx headlines matched.\n"
        f"window.__PUB_IMAGES__ = {js_data};\n",
        encoding='utf-8'
    )
    print(f"\nWrote {len(pub_images)} entries to {OUT_PATH}")


if __name__ == '__main__':
    main()
